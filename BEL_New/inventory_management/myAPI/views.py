from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
import json
from pymongo import MongoClient
import traceback
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
import secrets
import hashlib
import csv
from io import StringIO,BytesIO
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
import os
from io import StringIO
from datetime import datetime
from zoneinfo import ZoneInfo

client = MongoClient("mongodb://localhost:27017/")
db = client["inventory_db"]  # Use your DB name
collection = db["product_details"]
log_collection = db["api_logs"]
users_collection = db["users"]
sessions_collection = db["sessions"]
spares_master = db["spares_master"]   # master list
spares_in_col = db["spares_in"]       # logs table
spares_out_col = db["spares_out"]  # logs table
spares_audit = db["spares_audit"]

# Admin Projects collection
admin_projects_collection = db["admin_projects"]
# ----------------------
# Admin Projects Endpoints
# ----------------------

@csrf_exempt
def admin_add_project(request):
    """Add a new project (with duplicate check)"""
    user, err = require_auth(request, role="admin")
    if err:
        return err
    if request.method != "POST":
        return JsonResponse({"error": "Only POST allowed"}, status=405)
    try:
        body = json.loads(request.body or b"{}")
        project_name = (body.get("projectName") or "").strip()
        if not project_name:
            return JsonResponse({"error": "Project name required"}, status=400)
        exists = admin_projects_collection.find_one({"projectName": project_name})
        if exists:
            return JsonResponse({"error": "Project already exists"}, status=409)
        admin_projects_collection.insert_one({
            "projectName": project_name,
            "items": [],
            "createdBy": user.get("username"),
            "createdAt": datetime.now(ZoneInfo("Asia/Kolkata")),
        })
        return JsonResponse({"message": "Project created", "projectName": project_name}, status=201)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

@csrf_exempt
def admin_add_item(request):
    """Add item data to a project (with duplicate hierarchy check)"""
    user, err = require_auth(request, role="admin")
    if err:
        return err
    if request.method != "POST":
        return JsonResponse({"error": "Only POST allowed"}, status=405)
    try:
        body = json.loads(request.body or b"{}")
        project_name = (body.get("projectName") or "").strip()
        item_type = (body.get("itemType") or "").strip()
        item_name = (body.get("itemName") or "").strip()
        part_no = (body.get("partNo") or "").strip()
        if not (project_name and item_type and item_name and part_no):
            return JsonResponse({"error": "All fields required"}, status=400)
        # Check for duplicate in hierarchy
        exists = admin_projects_collection.find_one({
            "projectName": project_name,
            "items": {
                "$elemMatch": {
                    "partNo": part_no
                }
            }
        })
        if exists:
            return JsonResponse({"error": "Duplicate item in hierarchy"}, status=409)
        # Add item
        result = admin_projects_collection.update_one(
            {"projectName": project_name},
            {"$push": {"items": {
                "itemType": item_type,
                "itemName": item_name,
                "partNo": part_no,
                "createdBy": user.get("username"),
                "createdAt": datetime.now(ZoneInfo("Asia/Kolkata"))
            }}}
        )
        if result.matched_count == 0:
            return JsonResponse({"error": "Project not found"}, status=404)
        return JsonResponse({"message": "Item added"}, status=201)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

@csrf_exempt
def admin_edit_item(request):
    """Edit item data in a project"""
    user, err = require_auth(request, role="admin")
    if err:
        return err
    if request.method != "PUT":
        return JsonResponse({"error": "Only PUT allowed"}, status=405)
    try:
        body = json.loads(request.body or b"{}")
        project_name = (body.get("projectName") or "").strip()
        items = body.get("items")
        # If batch replace (Save All)
        if project_name and isinstance(items, list):
            result = admin_projects_collection.update_one(
                {"projectName": project_name},
                {"$set": {"items": items}}
            )
            if result.matched_count == 0:
                return JsonResponse({"error": "Project not found"}, status=404)
            return JsonResponse({"message": "All items replaced successfully", "count": len(items)})
        # Else, single item edit
        old_item_type = (body.get("oldItemType") or body.get("itemType") or "").strip()
        old_item_name = (body.get("oldItemName") or body.get("itemName") or "").strip()
        old_part_no = (body.get("oldPartNo") or body.get("partNo") or "").strip()
        new_data = body.get("newData") or {}
        if not (project_name and old_item_type and old_item_name and old_part_no and new_data):
            return JsonResponse({"error": "All fields required"}, status=400)
        result = admin_projects_collection.update_one(
            {"projectName": project_name, "items": {
                "$elemMatch": {
                    "itemType": old_item_type,
                    "itemName": old_item_name,
                    "partNo": old_part_no
                }
            }},
            {"$set": {"items.$": {
                **new_data,
                "updatedBy": user.get("username"),
                "updatedAt": datetime.now(ZoneInfo("Asia/Kolkata"))
            }}}
        )
        if result.matched_count == 0:
            return JsonResponse({"error": "Item not found"}, status=404)
        return JsonResponse({"message": "Item updated"})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

@csrf_exempt
def admin_delete_item(request):
    """Delete item data from a project"""
    user, err = require_auth(request, role="admin")
    if err:
        return err
    if request.method != "DELETE":
        return JsonResponse({"error": "Only DELETE allowed"}, status=405)
    try:
        body = json.loads(request.body or b"{}")
        project_name = (body.get("projectName") or "").strip()
        item_type = (body.get("itemType") or "").strip()
        item_name = (body.get("itemName") or "").strip()
        part_no = (body.get("partNo") or "").strip()
        if not (project_name and item_type and item_name and part_no):
            return JsonResponse({"error": "All fields required"}, status=400)
        result = admin_projects_collection.update_one(
            {"projectName": project_name},
            {"$pull": {"items": {
                "itemType": item_type,
                "itemName": item_name,
                "partNo": part_no
            }}}
        )
        if result.matched_count == 0:
            return JsonResponse({"error": "Project not found"}, status=404)
        return JsonResponse({"message": "Item deleted"})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

def admin_get_projects(request):
    """Get all projects"""
    user, err = require_auth(request)
    if err:
        return err
    if request.method != "GET":
        return JsonResponse({"error": "Only GET allowed"}, status=405)
    try:
        projects = list(admin_projects_collection.find({}, {"_id": 0, "projectName": 1}))
        return JsonResponse({"projects": [p["projectName"] for p in projects]})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

def admin_get_project_items(request):
    """Get all items for a project"""
    user, err = require_auth(request)
    if err:
        return err
    if request.method != "GET":
        return JsonResponse({"error": "Only GET allowed"}, status=405)
    try:
        project_name = request.GET.get("projectName", "").strip()
        if not project_name:
            return JsonResponse({"error": "Project name required"}, status=400)
        doc = admin_projects_collection.find_one({"projectName": project_name}, {"_id": 0, "items": 1})
        if not doc:
            return JsonResponse({"error": "Project not found"}, status=404)
        return JsonResponse({"items": doc.get("items", [])})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

# Bootstrap a default admin user if none exists
try:
    if users_collection.count_documents({}) == 0:
        users_collection.insert_one({
            "name": "Administrator",
            "username": "admin",
            "password_hash": hashlib.sha256(("bel_simple_salt" + "admin123").encode("utf-8")).hexdigest(),
            "role": "admin",
            "created_at": datetime.now(ZoneInfo("Asia/Kolkata"))
        })
except Exception:
    # If Mongo isn't reachable at import time, ignore; runtime endpoints will fail gracefully
    pass

# Clean up expired sessions (older than 24 hours)
def cleanup_expired_sessions():
    try:
        from datetime import timedelta
        cutoff_time = datetime.now(ZoneInfo("Asia/Kolkata")) - timedelta(hours=24)
        sessions_collection.delete_many({"created_at": {"$lt": cutoff_time}})
    except Exception:
        pass  # Ignore cleanup errors

# Logging function to store logs in MongoDB
def log_api_response(endpoint, method, request_data, response_data):
    log_entry = {
        "endpoint": endpoint,
        "method": method,
        "request_data": request_data,
        "response_data": response_data,
        "timestamp": datetime.now(ZoneInfo("Asia/Kolkata"))
    }
    log_collection.insert_one(log_entry)


# ----------------------
# Auth helpers
# ----------------------
def hash_password(password: str) -> str:
    salt = "bel_simple_salt"  # replace with env var in prod
    return hashlib.sha256((salt + password).encode("utf-8")).hexdigest()


def generate_token() -> str:
    return secrets.token_hex(32)


def get_auth_token_from_request(request):
    auth_header = request.headers.get("Authorization") or request.META.get("HTTP_AUTHORIZATION")
    if not auth_header:
        return None
    parts = auth_header.split()
    if len(parts) == 2 and parts[0].lower() == "bearer":
        return parts[1]
    return None


def get_user_from_token(request):
    token = get_auth_token_from_request(request)
    if not token:
        return None
    session = sessions_collection.find_one({"token": token})
    if not session:
        return None
    user = users_collection.find_one({"_id": session.get("user_id")})
    if not user:
        return None
    return {"id": str(user.get("_id")), "username": user.get("username"), "role": user.get("role"), "name": user.get("name")}


def require_auth(request, role: str | None = None):
    user = get_user_from_token(request)
    if not user:
        return None, JsonResponse({"error": "Unauthorized"}, status=401)
    if role and user.get("role") != role:
        return None, JsonResponse({"error": "Forbidden"}, status=403)
    return user, None


# ----------------------
# Auth endpoints
# ----------------------
@csrf_exempt
def login(request):
    if request.method != "POST":
        error_response = {"error": "Only POST allowed"}
        log_api_response("login", request.method, getattr(request, 'body', None), error_response)
        return JsonResponse(error_response, status=405)
    try:
        body = json.loads(request.body or b"{}")
        username = (body.get("username") or "").strip()
        password = body.get("password") or ""
        if not username or not password:
            response = {"error": "username and password are required"}
            log_api_response("login", request.method, body, response)
            return JsonResponse(response, status=400)

        user = users_collection.find_one({"username": username})
        if not user or user.get("password_hash") != hash_password(password):
            response = {"error": "Invalid credentials"}
            log_api_response("login", request.method, {"username": username}, response)
            return JsonResponse(response, status=401)

        # Clean up expired sessions before creating new one
        cleanup_expired_sessions()
        
        token = generate_token()
        sessions_collection.insert_one({
            "token": token,
            "user_id": user.get("_id"),
            "role": user.get("role"),
            "created_at": datetime.now(ZoneInfo("Asia/Kolkata"))
        })

        response = {
            "token": token, 
            "role": user.get("role"),
            "username": user.get("username"),
            "name": user.get("name")
        }
        log_api_response("login", request.method, {"username": username}, response)
        return JsonResponse(response)
    except Exception as e:
        stack_trace = traceback.format_exc()
        error_response = {"error": str(e)}
        log_api_response("login", request.method, getattr(request, 'body', None), {**error_response, "stack_trace": stack_trace})
        return JsonResponse(error_response, status=500)


@csrf_exempt
def validate_token(request):
    if request.method != "GET":
        error_response = {"error": "Only GET allowed"}
        log_api_response("validate_token", request.method, None, error_response)
        return JsonResponse(error_response, status=405)
    
    user, err = require_auth(request)
    if err:
        return err
    
    try:
        # Clean up expired sessions
        cleanup_expired_sessions()
        
        # If we get here, the token is valid
        response = {
            "valid": True,
            "username": user.get("username"),
            "role": user.get("role"),
            "name": user.get("name")
        }
        log_api_response("validate_token", request.method, None, response)
        return JsonResponse(response)
    except Exception as e:
        stack_trace = traceback.format_exc()
        error_response = {"error": str(e)}
        log_api_response("validate_token", request.method, None, {**error_response, "stack_trace": stack_trace})
        return JsonResponse(error_response, status=500)


@csrf_exempt
def logout(request):
    if request.method != "POST":
        error_response = {"error": "Only POST allowed"}
        log_api_response("logout", request.method, None, error_response)
        return JsonResponse(error_response, status=405)
    
    try:
        token = get_auth_token_from_request(request)
        if token:
            # Remove the session from the database
            sessions_collection.delete_one({"token": token})
        
        response = {"message": "Logged out successfully"}
        log_api_response("logout", request.method, None, response)
        return JsonResponse(response)
    except Exception as e:
        stack_trace = traceback.format_exc()
        error_response = {"error": str(e)}
        log_api_response("logout", request.method, None, {**error_response, "stack_trace": stack_trace})
        return JsonResponse(error_response, status=500)


@csrf_exempt
def admin_add_user(request):
    if request.method != "POST":
        error_response = {"error": "Only POST allowed"}
        log_api_response("admin_add_user", request.method, getattr(request, 'body', None), error_response)
        return JsonResponse(error_response, status=405)

    user, err = require_auth(request, role="admin")
    if err:
        return err
    try:
        body = json.loads(request.body or b"{}")
        name = (body.get("name") or "").strip()
        username = (body.get("username") or "").strip().lower()
        password = body.get("password") or ""
        role = (body.get("role") or "user").strip().lower()

        if not name or not username or not password:
            response = {"error": "name, username, password are required"}
            log_api_response("admin_add_user", request.method, body, response)
            return JsonResponse(response, status=400)

        if role not in ["admin", "user"]:
            role = "user"

        exists = users_collection.find_one({"username": username})
        if exists:
            response = {"error": "username already exists"}
            log_api_response("admin_add_user", request.method, {"username": username}, response)
            return JsonResponse(response, status=409)

        doc = {
            "name": name,
            "username": username,
            "password_hash": hash_password(password),
            "role": role,
            "created_at": datetime.now(ZoneInfo("Asia/Kolkata"))
        }
        users_collection.insert_one(doc)
        response = {"message": "user created", "username": username, "role": role}
        log_api_response("admin_add_user", request.method, {"admin": user.get("username"), "new_user": username}, response)
        return JsonResponse(response, status=201)
    except Exception as e:
        stack_trace = traceback.format_exc()
        error_response = {"error": str(e)}
        log_api_response("admin_add_user", request.method, getattr(request, 'body', None), {**error_response, "stack_trace": stack_trace})
        return JsonResponse(error_response, status=500)
 


# ----------------------
# Inventory: Item In / Out / CRUD by passNo
# ----------------------
@csrf_exempt
def items_in(request):
    if request.method != "POST":
        error_response = {"error": "Only POST allowed"}
        log_api_response("items_in", request.method, getattr(request, 'body', None), error_response)
        return JsonResponse(error_response, status=405)
    user, err = require_auth(request)
    if err:
        return err
    try:
        body = json.loads(request.body or b"{}")
        pass_no = (body.get("passNo") or "").strip()
        date_in = (body.get("dateIn") or datetime.now(ZoneInfo("Asia/Kolkata")).date().isoformat())
        customer = {
            "name": body.get("customerName"),
            "unitAddress": body.get("customerUnitAddress"),
            "location": body.get("customerLocation"),
            "phone": body.get("customerPhoneNo"),
        }
        project_name = body.get("projectName")
        items = body.get("items") or []

        if not pass_no:
            response = {"error": "passNo is required"}
            log_api_response("items_in", request.method, body, response)
            return JsonResponse(response, status=400)

        exists = collection.find_one({"passNo": pass_no})
        if exists:
            response = {"error": "passNo already exists"}
            log_api_response("items_in", request.method, {"passNo": pass_no}, response)
            return JsonResponse(response, status=409)

        normalized_items = []
        for it in items:
            normalized_items.append({
                "equipmentType": it.get("equipmentType"),
                "itemName": it.get("itemName"),
                "partNumber": it.get("partNumber"),
                "serialNumber": it.get("serialNumber"),
                "defectDetails": it.get("defectDetails"),
                "itemIn": True,  # Always true when item is entered
                "itemOut": False,
                "dateOut": None,  # Will be set when item goes out
                "itemRectificationDetails": "",  # New field for rectification details
                "itemFeedback1Details": "",  # New field for feedback 1 details
                "itemFeedback2Details": "",  # New field for feedback 2 details
            })
        
        # Sort items by part number in ascending order
        normalized_items.sort(key=lambda x: x.get("partNumber", ""))

        doc = {
            "passNo": pass_no,
            "dateIn": date_in,
            "customer": customer,
            "projectName": project_name,
            "items": normalized_items,
            "createdBy": user.get("username"),
            "createdAt": datetime.now(ZoneInfo("Asia/Kolkata")),
            "updatedAt": datetime.now(ZoneInfo("Asia/Kolkata")),
        }
        collection.insert_one(doc)
        doc.pop("_id", None)
        response = {"message": "Item In recorded", "data": doc}
        log_api_response("items_in", request.method, {"passNo": pass_no}, response)
        return JsonResponse(response, status=201)
    except Exception as e:
        stack_trace = traceback.format_exc()
        error_response = {"error": str(e)}
        log_api_response("items_in", request.method, getattr(request, 'body', None), {**error_response, "stack_trace": stack_trace})
        return JsonResponse(error_response, status=500)

def get_item_by_passno(request, pass_no):
    if not pass_no:
        print("Enter a pass number")  # <--- your print statement
        error_response = {"error": "passNo cannot be empty"}
        log_api_response("get_item_by_passno", request.method, {"passNo": pass_no}, error_response)
        return JsonResponse(error_response, status=400)
    
    if request.method != "GET":
        error_response = {"error": "Only GET allowed"}
        log_api_response("get_item_by_passno", request.method, {"passNo": pass_no}, error_response)
        return JsonResponse(error_response, status=405)
    user, err = require_auth(request)
    if err:
        return err
    try:
        doc = collection.find_one({"passNo": pass_no}, {"_id": 0})
        if not doc:
            response = {"error": "Not found"}
            log_api_response("get_item_by_passno", request.method, {"passNo": pass_no}, response)
            return JsonResponse(response, status=404)
        
        log_api_response("get_item_by_passno", request.method, {"passNo": pass_no}, doc)
        return JsonResponse(doc, safe=False)
    except Exception as e:
        stack_trace = traceback.format_exc()
        error_response = {"error": str(e)}
        log_api_response("get_item_by_passno", request.method, {"passNo": pass_no}, {**error_response, "stack_trace": stack_trace})
        return JsonResponse(error_response, status=500)

@csrf_exempt
def update_item_rfd(request, pass_no):
    if request.method != "PUT":
        error_response = {"error": "Only PUT allowed"}
        log_api_response("update_item_rfd", request.method, {"passNo": pass_no}, error_response)
        return JsonResponse(error_response, status=405)
    user, err = require_auth(request)
    if err:
        return err
    try:
        print(f"=== ITEM RFD UPDATE DEBUG ===")
        print(f"Pass Number: {pass_no}")
        print(f"User: {user.get('username')}")
        print(f"Request body: {request.body}")

        body = json.loads(request.body or b"{}")
        updates = body.get("items") or []

        print(f"Parsed body: {body}")
        print(f"Updates array: {updates}")

        doc = collection.find_one({"passNo": pass_no})
        if not doc:
            response = {"error": "Not found"}
            log_api_response("update_item_rfd", request.method, {"passNo": pass_no}, response)
            return JsonResponse(response, status=404)
        
        print(f"Found document: {doc.get('passNo')}")
        print(f"Original items: {doc.get('items')}")

        # Check if we have the same number of items
        original_items = doc.get("items", [])
        if len(updates) != len(original_items):
            response = {"error": f"Number of items mismatch. Expected {len(original_items)}, got {len(updates)}"}
            return JsonResponse(response, status=400)

        print(f"DEBUG: Received updates: {updates}")

        # Update items by position (index) instead of serial number
        new_items = []
        for i, (original_item, update_item) in enumerate(zip(original_items, updates)):
            print(f"Processing item {i}: original={original_item.get('serialNumber')}, update={update_item.get('serialNumber')}")

            # Create updated item
            updated_item = original_item.copy()
            updated_item["itemRfd"] = bool(update_item.get("itemRfd", False))

            # Prevent RFD being unset for itemOut items
            if original_item.get("itemOut") is True:
                updated_item["itemRfd"] = True

            # Handle dateRfd
            if updated_item["itemRfd"]:
                if update_item.get("dateRfd"):
                    updated_item["dateRfd"] = update_item["dateRfd"]
                elif not original_item.get("dateRfd"):
                    updated_item["dateRfd"] = datetime.now(ZoneInfo("Asia/Kolkata")).date().isoformat()
                    print(f"DEBUG: Auto-setting dateRfd for item {i} to {updated_item['dateRfd']}")
            else:
                updated_item["dateRfd"] = None
                print(f"DEBUG: Clearing dateRfd for item {i} since itemRfd is False")

            # Handle rectification details
            if "itemRectificationDetails" in update_item:
                updated_item["itemRectificationDetails"] = update_item["itemRectificationDetails"] or ""

            if "itemFeedback1Details" in update_item:
                updated_item["itemFeedback1Details"] = update_item["itemFeedback1Details"] or ""

            if "itemFeedback2Details" in update_item:
                updated_item["itemFeedback2Details"] = update_item["itemFeedback2Details"] or ""
            
            print(f"Updated item {i}: {updated_item}")
            new_items.append(updated_item)    

        print(f"Final items array: {new_items}")

        result = collection.update_one({"passNo": pass_no}, {"$set": {"items": new_items, "updatedAt": datetime.now(ZoneInfo("Asia/Kolkata")), "updatedBy": user.get("username")}})
        print(f"Update result: matched={result.matched_count}, modified={result.modified_count}")

        response = {"message": "RFD statuses updated"}
        log_api_response("update_item_rfd", request.method, {"passNo": pass_no, "updates_count": len(updates)}, response)
        return JsonResponse(response)
    except Exception as e:
        print(f"ERROR in update_item_rfd: {str(e)}")
        print(f"Traceback: {traceback.format_exc()}")
        stack_trace = traceback.format_exc()
        error_response = {"error": str(e)}
        log_api_response("update_item_rfd", request.method, {"passNo": pass_no}, {**error_response, "stack_trace": stack_trace})
        return JsonResponse(error_response, status=500)

@csrf_exempt
def update_item_out(request, pass_no):
    if request.method != "PUT":
        error_response = {"error": "Only PUT allowed"}
        log_api_response("update_item_out", request.method, {"passNo": pass_no}, error_response)
        return JsonResponse(error_response, status=405)
    user, err = require_auth(request)
    if err:
        return err
    try:
        print(f"=== ITEM OUT UPDATE DEBUG ===")
        print(f"Pass Number: {pass_no}")
        print(f"User: {user.get('username')}")
        print(f"Request body: {request.body}")
        
        body = json.loads(request.body or b"{}")
        updates = body.get("items") or []

        print(f"Parsed body: {body}")
        print(f"Updates array: {updates}")

        doc = collection.find_one({"passNo": pass_no})
        if not doc:
            response = {"error": "Not found"}
            log_api_response("update_item_out", request.method, {"passNo": pass_no}, response)
            return JsonResponse(response, status=404)

        print(f"Found document: {doc.get('passNo')}")
        print(f"Original items: {doc.get('items')}")

        # Check if we have the same number of items
        original_items = doc.get("items", [])
        if len(updates) != len(original_items):
            response = {"error": f"Number of items mismatch. Expected {len(original_items)}, got {len(updates)}"}
            return JsonResponse(response, status=400)

        print(f"DEBUG: Received updates: {updates}")
        
        # Update items by position (index) instead of serial number
        new_items = []
        for i, (original_item, update_item) in enumerate(zip(original_items, updates)):
            print(f"Processing item {i}: original={original_item.get('serialNumber')}, update={update_item.get('serialNumber')}")
            
            # Create updated item
            updated_item = original_item.copy()
            updated_item["itemOut"] = bool(update_item.get("itemOut", False))
            
            if not original_item.get("itemRfd"):
                updated_item["itemOut"] = False

            if original_item.get("itemOut") is True:
                updated_item["itemOut"] = True

            # Handle dateOut
            if updated_item["itemOut"]:
                if update_item.get("dateOut"):
                    updated_item["dateOut"] = update_item["dateOut"]
                elif not original_item.get("dateOut"):
                    updated_item["dateOut"] = datetime.now(ZoneInfo("Asia/Kolkata")).date().isoformat()
                    print(f"DEBUG: Auto-setting dateOut for item {i} to {updated_item['dateOut']}")
            else:
                updated_item["dateOut"] = None
                print(f"DEBUG: Clearing dateOut for item {i} since itemOut is False")
            
            # Handle rectification details
            if "itemRectificationDetails" in update_item:
                updated_item["itemRectificationDetails"] = update_item["itemRectificationDetails"] or ""

            if "itemFeedback1Details" in update_item:
                updated_item["itemFeedback1Details"] = update_item["itemFeedback1Details"] or ""

            if "itemFeedback2Details" in update_item:
                updated_item["itemFeedback2Details"] = update_item["itemFeedback2Details"] or ""
            
            print(f"Updated item {i}: {updated_item}")
            new_items.append(updated_item)

        print(f"Final items array: {new_items}")
        
        result = collection.update_one({"passNo": pass_no}, {"$set": {"items": new_items, "updatedAt": datetime.now(ZoneInfo("Asia/Kolkata")), "updatedBy": user.get("username")}})
        print(f"Update result: matched={result.matched_count}, modified={result.modified_count}")
        
        response = {"message": "ItemOut statuses updated"}
        log_api_response("update_item_out", request.method, {"passNo": pass_no, "updates_count": len(updates)}, response)
        return JsonResponse(response)
    except Exception as e:
        print(f"ERROR in update_item_out: {str(e)}")
        print(f"Traceback: {traceback.format_exc()}")
        stack_trace = traceback.format_exc()
        error_response = {"error": str(e)}
        log_api_response("update_item_out", request.method, {"passNo": pass_no}, {**error_response, "stack_trace": stack_trace})
        return JsonResponse(error_response, status=500)

@csrf_exempt
def edit_record(request, pass_no):
    user, err = require_auth(request)
    if err:
        return err
    try:
        if request.method == "GET":
            doc = collection.find_one({"passNo": pass_no}, {"_id": 0})
            if not doc:
                response = {"error": "Entry Not found"}
                log_api_response("edit_record", request.method, {"passNo": pass_no}, response)
                return JsonResponse(response, status=404)
            log_api_response("edit_record", request.method, {"passNo": pass_no}, doc)
            return JsonResponse(doc, safe=False)
        elif request.method == "PUT":
            body = json.loads(request.body or b"{}")
            if body.get("passNo") and body.get("passNo") != pass_no:
                response = {"error": "passNo cannot be changed"}
                log_api_response("edit_record", request.method, {"passNo": pass_no}, response)
                return JsonResponse(response, status=400)

            allowed_fields = ["dateIn", "customer", "projectName", "items"]
            set_fields = {k: v for k, v in body.items() if k in allowed_fields}
            
            # Sort items by part number if items are being updated
            if "items" in set_fields:
                for item in set_fields["items"]:
                    item["itemIn"] = True  # Always true when item is entered
                    if "dateOut" not in item:
                        item["dateOut"] = None
                    if "dateRfd" not in item:
                        item["dateRfd"] = None
                    if "itemRectificationDetails" not in item:
                        item["itemRectificationDetails"] = ""
                    if "itemFeedback1Details" not in item:
                        item["itemFeedback1Details"] = ""
                    if "itemFeedback2Details" not in item:
                        item["itemFeedback2Details"] = ""
                set_fields["items"].sort(key=lambda x: x.get("partNumber", ""))
            
            set_fields["updatedAt"] = datetime.now(ZoneInfo("Asia/Kolkata"))
            set_fields["updatedBy"] = user.get("username")
            result = collection.update_one({"passNo": pass_no}, {"$set": set_fields})
            if result.matched_count == 0:
                response = {"error": "Not found"}
                log_api_response("edit_record", request.method, {"passNo": pass_no}, response)
                return JsonResponse(response, status=404)
            response = {"message": "Record updated"}
            log_api_response("edit_record", request.method, {"passNo": pass_no}, response)
            return JsonResponse(response)
        elif request.method == "DELETE":
            result = collection.delete_one({"passNo": pass_no})
            if result.deleted_count == 0:
                response = {"error": "Not found"}
                log_api_response("edit_record", request.method, {"passNo": pass_no}, response)
                return JsonResponse(response, status=404)
            response = {"message": "Record deleted"}
            log_api_response("edit_record", request.method, {"passNo": pass_no}, response)
            return JsonResponse(response)
        else:
            error_response = {"error": "Method not allowed"}
            log_api_response("edit_record", request.method, {"passNo": pass_no}, error_response)
            return JsonResponse(error_response, status=405)
    except Exception as e:
        stack_trace = traceback.format_exc()
        error_response = {"error": str(e)}
        log_api_response("edit_record", request.method, {"passNo": pass_no}, {**error_response, "stack_trace": stack_trace})
        return JsonResponse(error_response, status=500)


# ----------------------
# Search + download
# ----------------------
def _build_date_filter(from_str: str | None, to_str: str | None):
    if not from_str and not to_str:
        return None
    cond = {}
    if from_str:
        cond["$gte"] = from_str
    if to_str:
        cond["$lte"] = to_str
    return cond


def _build_search_query(params):
    search_type = params.get("type")
    value = params.get("value")
    from_date = params.get("from")
    to_date = params.get("to")
    serialProjectName = params.get("serialProjectName")
    query = {}
    if search_type == "passNo" and value:
        query["passNo"] = value
    elif search_type == "serialNumber" and value:
        if serialProjectName:
            query["projectName"] = {"$regex": serialProjectName, "$options": "i"}    
        query["items.serialNumber"] = {"$regex": value.upper(), "$options": "i"}
    elif search_type == "ItemPartNo" and value:
        query["items.partNumber"] = value
    elif search_type == "ProjectName" and value:
        query["projectName"] = {"$regex": value, "$options": "i"}
    elif search_type == "DateRange":
        pass  # only date filter
    date_cond = _build_date_filter(from_date, to_date)
    if date_cond:
        query["dateIn"] = date_cond
    print(query)
    return query

def _filter_serial(items, serial_substring=None, status=None):
    """Filter items by serial number substring (case-insensitive) and status."""
    if not serial_substring:
        return items

    serial_substring = serial_substring.upper()
    filtered = []

    for item in items:
        serial = item.get("serialNumber", "")

        # 🔹 Serial number substring match
        if not (serial and serial_substring in serial.upper()):
            continue

        # 🔹 Status filter (optional)
        if status == "In" and not (item.get("itemIn") and not item.get("itemOut") and not item.get("itemRfd")):
            continue
        if status == "RFD" and not (item.get("itemIn") and item.get("itemRfd") and not item.get("itemOut")):
            continue
        if status == "Out" and not (item.get("itemIn") and item.get("itemOut")):
            continue

        filtered.append(item)

    return filtered

def _filter_items(items, part_number=None, status=None):
    """Filter items by part number and/or status (In/Out)."""
    filtered = []
    for item in items:
        # Part number filter
        if part_number and item.get("partNumber") != part_number:
            continue

        # Status filter
        if status == "In" and not (item.get("itemIn") and not item.get("itemOut") and not item.get("itemRfd")):
            continue

        if status == "RFD" and not (item.get("itemIn") and item.get("itemRfd") and not item.get("itemOut")):
            continue

        if status == "Out" and not (item.get("itemIn") and item.get("itemOut")):
            continue

        filtered.append(item)
    return filtered


def _filter_items_by_part_number(items, part_number):
    """Filter items to only include those matching the part number search"""
    if not part_number:
        return items
    return [item for item in items if item.get("partNumber") == part_number]


def _shape_search_result(doc):
    return {
        "passNo": doc.get("passNo"),
        "projectName": doc.get("projectName"),
        "dateIn": doc.get("dateIn"),
        "customer": doc.get("customer", {}),
        "items": doc.get("items", []),
        "createdBy": doc.get("createdBy", ""),
        "updatedBy": doc.get("updatedBy", ""),
    }

def search(request):
    if request.method != "GET":
        error_response = {"error": "Only GET allowed"}
        log_api_response("search", request.method, dict(request.GET), error_response)
        return JsonResponse(error_response, status=405)

    user, err = require_auth(request)
    if err:
        return err

    try:
        params = request.GET
        query = _build_search_query(params)
        docs = list(collection.find(query))

        results = []
        serial_no = 1
        search_type = params.get("type")
        search_value = params.get("value")
        status = params.get("status")  # "In" or "Out"

        for doc in docs:
            filtered_items = doc.get("items", [])

            if search_type == "serialNumber" and search_value:
                filtered_items = _filter_serial(filtered_items, serial_substring=search_value, status=status)
            elif search_type == "ItemPartNo" and search_value:
                filtered_items = _filter_items(filtered_items, part_number=search_value, status=status)
            elif status in ("In", "RFD", "Out"):
                filtered_items = _filter_items(filtered_items, status=status)

            for item in filtered_items:
                item["serialNo"] = serial_no
                serial_no += 1

            doc = {**doc, "items": filtered_items}
            results.append(_shape_search_result({**doc, "_id": None}))

        response = {"count": len(results), "data": results}
        log_api_response("search", request.method, dict(params), {"count": response["count"]})
        return JsonResponse(response)

    except Exception as e:
        stack_trace = traceback.format_exc()
        error_response = {"error": str(e)}
        log_api_response("search", request.method, dict(request.GET), {**error_response, "stack_trace": stack_trace})
        return JsonResponse(error_response, status=500)

@csrf_exempt
def search_download(request):
    if request.method != "GET":
        error_response = {"error": "Only GET allowed"}
        log_api_response("search_download", request.method, dict(request.GET), error_response)
        return JsonResponse(error_response, status=405)

    user, err = require_auth(request)
    if err:
        return err

    try:
        serial_no = 1
        params = request.GET
        query = _build_search_query(params)
        docs = list(collection.find(query))

        # Create CSV content
        output = StringIO()
        writer = csv.writer(output)
        
        # Write header row
        writer.writerow([
            "Sl No.","Pass No", "Project Name", 
            "Customer Name", "Customer Unit Address", "Customer Location", "Customer Phone",
            "Equipment Type", "Item Name", "Part Number", "Serial Number", "Defect Details", 
            "Status", "Date In", "Date RFD", "Date Out", "Item Rectification Details", "Feedback 1 details", "Feedback 2 details", "CreatedBy", "updatedBy"
        ])
        
        # Write data rows - one row per item
        for doc in docs:
            pass_no = doc.get("passNo", "")
            date_in = doc.get("dateIn", "")
            project_name = doc.get("projectName", "")
            customer = doc.get("customer", {})
            items = doc.get("items", [])
            createdBy = doc.get("createdBy", "")
            updatedBy = doc.get("updatedBy", "")
            
            search_type = params.get("type")
            status = params.get("status")
            search_value = params.get("value")
            if search_type == "serialNumber" and search_value:
                items = _filter_serial(items, serial_substring=search_value, status=status)
            elif search_type == "ItemPartNo" and search_value:
                items = _filter_items(items, part_number = search_value, status=status)
            elif status in ("In", "RFD", "Out"):
                items = _filter_items(items, status=status)

            # Filter items by part number if searching by part number
            
            # if search_type == "ItemPartNo" and params.get("value"):
            #     items = _filter_items_by_part_number(items, params.get("value"))
            
            for item in items:
                # Determine status: OUT if both itemIn and itemOut are true, else IN
                status = "OUT" if item.get("itemIn") and item.get("itemOut") else "IN"
                status = "RFD" if item.get("itemIn") and item.get("itemRfd") and not item.get("itemOut") else status
                
                # Format phone number properly (remove scientific notation)
                phone = customer.get("phone", "")
                if phone and str(phone).isdigit():
                    phone = str(phone)
                
                # Format date properly for Excel
                date_rfd = item.get("dateRfd", "")
                if date_rfd:
                    # Ensure date is in YYYY-MM-DD format
                    try:
                        if isinstance(date_rfd, str):
                            date_rfd = date_rfd[:10]  # Take first 10 characters
                    except:
                        date_rfd = ""

                date_out = item.get("dateOut", "")
                if date_out:
                    # Ensure date is in YYYY-MM-DD format
                    try:
                        if isinstance(date_out, str):
                            date_out = date_out[:10]  # Take first 10 characters
                    except:
                        date_out = ""

                writer.writerow([
                    serial_no,
                    pass_no,
                    project_name,
                    customer.get("name", ""),
                    customer.get("unitAddress", ""),
                    customer.get("location", ""),
                    phone,
                    item.get("equipmentType", ""),
                    item.get("itemName", ""),
                    item.get("partNumber", ""),
                    item.get("serialNumber", ""),
                    item.get("defectDetails", ""),
                    status,
                    date_in,
                    date_rfd,
                    date_out,
                    item.get("itemRectificationDetails", ""),
                    item.get("itemFeedback1Details", ""),
                    item.get("itemFeedback2Details", ""),
                    createdBy,
                    updatedBy
                ])
                serial_no += 1

        csv_content = output.getvalue()
        output.close()
        
        default_filename = f"{datetime.now(ZoneInfo('Asia/Kolkata')).strftime('%Y-%m-%d')}_inventory_export.csv"
        # Return CSV file
        response = HttpResponse(csv_content, content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{default_filename}"'
        
        log_api_response("search_download", request.method, dict(params), {"rows": len(docs)})
        return response

    except Exception as e:
        stack_trace = traceback.format_exc()
        error_response = {"error": str(e)}
        log_api_response("search_download", request.method, dict(request.GET), {**error_response, "stack_trace": stack_trace})
        return JsonResponse(error_response, status=500)

@csrf_exempt
def search_download_sticker(request):
    if request.method != "GET":
        error_response = {"error": "Only GET allowed"}
        log_api_response("search_download_sticker", request.method, dict(request.GET), error_response)
        return JsonResponse(error_response, status=405)

    user, err = require_auth(request)
    if err:
        return err
    
    try:
        BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        master_file = os.path.join(BASE_DIR, "static", "templates", "Print_Pass_Master_Excel.xlsx")
        wb = load_workbook(master_file)
        ws = wb.active
        params = request.GET
        poffset = int(params.get("offset","0"))
        query = _build_search_query(params)
        docs = list(collection.find(query))
        for doc in docs:
            passNo = doc.get("passNo","")
            dateIn = doc.get("dateIn","")
            items = doc.get("items", [])
            projectName = doc.get("projectName","")
            unitAddress = doc.get("customer",{}).get("unitAddress","")
            for j, item in enumerate(items):
                start_slNo = j+1
                itemName = item.get("itemName", "")
                serialNumber = item.get("serialNumber", "")
                j += poffset-1
                # Choose left or right block based on index
                if j % 2 == 0:
                    col_A, col_B, col_C = "A", "B", "C"
                else:
                    col_A, col_B, col_C = "D", "E", "F"

                start_row = (j // 2) * 7 + 1  # ensures each block gets 7 rows
                if start_row in (57,112):
                    start_row = (j//2) * 7
                    print(start_row)

                labels = ["Pass No", "Date", "Unit Address", "Project Name", "Item Name", "Sl.No"]
                values = [passNo, dateIn, unitAddress, projectName, itemName, serialNumber]

                # write serial number in first column
                ws[f"{col_A}{start_row}"] = start_slNo

                # write labels and values
                for offset, (label, value) in enumerate(zip(labels, values)):
                    ws[f"{col_B}{start_row + offset}"] = label
                    ws[f"{col_C}{start_row + offset}"] = value

                # empty row separator
                empty_row = start_row + len(labels)
                if empty_row != 56:
                    ws[f"{col_B}{empty_row}"] = ""

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        default_filename = f"{datetime.now(ZoneInfo('Asia/Kolkata')).strftime('%Y-%m-%d')}_sticker_export.xlsx"
        response = HttpResponse(output.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = f'attachment; filename="{default_filename}"'
        log_api_response("search_download_sticker", request.method, dict(params), {"rows": len(docs)})
        return response
    except Exception as e:
        stack_trace = traceback.format_exc()
        error_response = {"error": str(e)}
        log_api_response("search_download_sticker", request.method, dict(request.GET), {**error_response, "stack_trace": stack_trace})
        return JsonResponse(error_response, status=500)

@csrf_exempt
def search_download_form(request):
    if request.method != "GET":
        return JsonResponse({"error": "Only GET allowed"}, status=405)

    user, err = require_auth(request)
    if err:
        return err

    try:
        # --- Extract part numbers from query params ---
        part_numbers_str = request.GET.get("PartNumbers", "")
        allowed_part_numbers = set(p.strip() for p in part_numbers_str.split(",") if p.strip())

        query = _build_search_query(request.GET)
        docs = list(collection.find(query))
        if not docs:
            return JsonResponse({"error": "No documents found"}, status=404)

        # ---------- EXCEL SETUP ----------
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            std = wb["Sheet"]
            wb.remove(std)

        # page_number = 1
        # first_doc = docs[0] if docs else {}
        # ws = create_page(wb, page_number, first_doc)
        
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        MAX_ITEMS_PER_PAGE = 10  # Items per page

        # Helper functions
        def style_cell(cell, bold=False, size=10, align="left", wrap=False):
            cell.font = Font(bold=bold, size=size)
            cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)

        label_rows = {
            3: [("PASS NO:", "A", "B"), ("DATE:", "E", "E")],
            4: [("PASS DATE:", "A", "B"), ("CUSTOMER NAME:", "E", "E")],
            5: [("PROJECT NAME:", "A", "B"), ("CUSTOMER CONTACT NO:", "E", "E")],
            6: [("UNIT ADDRESS:", "A", "B"), ("CUSTOMER LOCATION:", "E", "E")]
        }

        headers = ["SL. NO", "PART NO", "ITEM NAME", "ITEM S1.N",
                   "DEFECT DETAILS", "RECTIFICATION DETAILS", "REMARKS"]

        column_widths = {'A': 5, 'B': 20, 'C': 25, 'D': 15, 'E': 20, 'F': 40, 'G': 10}

        # ---------- FUNCTIONS ----------
        def create_page(wb, page_number, doc):
            ws = wb.create_sheet(title=f"Customer Support MILCOM - Page {page_number}")
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            for i in range(3, 7):
                ws.row_dimensions[i].height = 10
            ws.row_dimensions[7].height = 5

            # Titles
            ws.merge_cells('A1:G1')
            ws['A1'] = "CUSTOMER SUPPORT MILCOM"
            style_cell(ws['A1'], bold=True, size=16, align="center")

            ws.merge_cells('A2:G2')
            ws['A2'] = "Customer Complaint History Card"
            style_cell(ws['A2'], bold=True, size=12, align="center")

            # Header Labels
            cust = doc.get("customer", {})
            header_values = {
                'C3:D3': doc.get("passNo", ""),
                'F3:G3': datetime.now(ZoneInfo("Asia/Kolkata")).strftime("%d-%m-%Y"),
                'C4:D4': doc.get("dateIn", ""),
                'F4:G4': cust.get("name", ""),
                'C5:D5': doc.get("projectName", ""),
                'F5:G5': cust.get("phone", ""),
                'C6:D6': cust.get("unitAddress", ""),
                'F6:G6': cust.get("location", "")
            }

            for row, pairs in label_rows.items():
                for label, start_col, end_col in pairs:
                    if start_col != end_col:
                        ws.merge_cells(f"{start_col}{row}:{end_col}{row}")
                    c = ws[f"{start_col}{row}"]
                    c.value = label
                    style_cell(c, bold=True, size=8, align="left")

            for cells, value in header_values.items():
                ws.merge_cells(cells)
                c = ws[cells.split(":")[0]]
                c.value = value
                style_cell(c, align="left")

            # Table header
            for col, h in enumerate(headers, start=1):
                c = ws.cell(row=8, column=col, value=h)
                style_cell(c, bold=True, size=9, align="center", wrap=True)
                c.border = thin_border

            # --- Set minimum row height for item rows ---
            for r in range(9, 19):
                ws.row_dimensions[r].height = 32

            return ws

        def create_footer(ws, start_row=19, end_row=23):
            footers = ["HANDED OVER BY (CS-Rep)", "RECEIVED BY (TS-Rep)", "RECEIVED BACK BY AFTER REPAIR (CS-Rep)"]
            ws.merge_cells(f'A{start_row}:B{start_row}')
            ws.merge_cells(f'C{start_row}:D{start_row}')
            ws.merge_cells(f'E{start_row}:G{start_row}')
            ws['A19'], ws['C19'], ws['E19'] = footers
            for cell in ['A19', 'C19', 'E19']:
                style_cell(ws[cell], bold=True)
                ws[cell].alignment = Alignment(horizontal="center", vertical="center")

            for row in range(start_row + 1, end_row + 1):
                for col in range(1, 8):
                    c = ws.cell(row=row, column=col)
                    c.border = thin_border
                    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

            ws.merge_cells(f'A{start_row+1}:B{end_row}')
            ws.merge_cells(f'C{start_row+1}:D{end_row}')
            ws.merge_cells(f'E{start_row+1}:G{end_row}')

        # ---------- POPULATE ----------
        serial_number = 1
        current_row = 9
        page_number = 1
        first_doc = docs[0] if docs else {}
        ws = create_page(wb, page_number, first_doc)

        for doc in docs:
            items = [i for i in doc.get("items", []) if str(i.get("partNumber", "")).strip() in allowed_part_numbers]
            for item in items:
                if (current_row - 9) % MAX_ITEMS_PER_PAGE == 0 and current_row != 9:
                    create_footer(ws)
                    page_number += 1
                    ws = create_page(wb, page_number, doc)
                    current_row = 9

                row_values = [
                    serial_number, item.get("partNumber", ""), item.get("itemName", ""),
                    item.get("serialNumber", ""), item.get("defectDetails", ""),
                    "", ""
                ]
                for col, val in enumerate(row_values, start=1):
                    c = ws.cell(row=current_row, column=col, value=val)
                    c.border = thin_border
                    align = "center" if col == 1 else "left"
                    c.alignment = Alignment(horizontal=align, vertical="top", wrap_text=True)
                serial_number += 1
                current_row += 1
            if ws:
                create_footer(ws)

        # ---------- RESPONSE ----------
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename="Customer_Complaint_History_Card.xlsx"'
        return response

    except Exception as e:
        stack_trace = traceback.format_exc()
        return JsonResponse({"error": str(e), "stack": stack_trace}, status=500)
    
@csrf_exempt
def search_suggestions(request):
    if request.method != "GET":
        error_response = {"error": "Only GET allowed"}
        log_api_response("search_suggestions", request.method, dict(request.GET), error_response)
        return JsonResponse(error_response, status=405)

    user, err = require_auth(request)
    if err:
        return err

    try:
        params = request.GET
        search_type = params.get("type")
        value = (params.get("value") or "").strip()
        if not search_type or not value:
            response = {"error": "type and value are required"}
            log_api_response("search_suggestions", request.method, dict(params), response)
            return JsonResponse(response, status=400)

        suggestions = set()
        if search_type == "passNo":
            docs = collection.find({"passNo": {"$regex": f"^{value}", "$options": "i"}}, {"_id": 0, "passNo": 1}).limit(10)
            for doc in docs:
                suggestions.add(doc.get("passNo"))
        elif search_type == "ItemPartNo":
            docs = collection.find({"items.partNumber": {"$regex": f"^{value}", "$options": "i"}}, {"_id": 0, "items.partNumber": 1}).limit(50)
            for doc in docs:
                for item in doc.get("items", []):
                    part_no = item.get("partNumber")
                    if part_no and part_no.lower().startswith(value.lower()):
                        suggestions.add(part_no)
                        if len(suggestions) >= 10:
                            break
                if len(suggestions) >= 10:
                    break
        elif search_type == "ProjectName":
            docs = admin_projects_collection.find({"projectName": {"$regex": f"^{value}", "$options": "i"}}, {"_id": 0, "projectName": 1}).limit(10)
            for doc in docs:
                suggestions.add(doc.get("projectName"))
        else:
            response = {"error": "Invalid type"}
            log_api_response("search_suggestions", request.method, dict(params), response)
            return JsonResponse(response, status=400)

        suggestions = list(suggestions)[:10]  # Limit to 10 suggestions
        response = {"suggestions": suggestions}
        log_api_response("search_suggestions", request.method, dict(params), {"count": len(suggestions)})
        return JsonResponse(response)
    except Exception as e:
        stack_trace = traceback.format_exc()
        error_response = {"error": str(e)}
        log_api_response("search_suggestions", request.method, dict(request.GET), {**error_response, "stack_trace": stack_trace})
        return JsonResponse(error_response, status=500)

@csrf_exempt
def spares_master_add(request):
    if request.method != "POST":
        return JsonResponse({"error": "Only POST allowed"}, status=405)

    user, err = require_auth(request)
    if err:
        return err

    try:
        body = json.loads(request.body or "{}")

        part_no = body.get("part_no", "").strip()
        item_name = body.get("item_name", "").strip()
        no_of_bins = body.get("no_of_bins")
        bin_nos = body.get("bin_nos", [])
        rack_no = body.get("rack_no", "").strip()
        item_loc = body.get("item_loc", "").strip()

        if not isinstance(bin_nos, list) or len(bin_nos) != int(no_of_bins):
            return JsonResponse(
                {"error": "bin_nos count must match no_of_bins"},
                status=400
            )
        # Required fields
        if not part_no or not item_name:
            return JsonResponse({"error": "part_no and item_name are required"}, status=400)

        # Connect to MongoDB
        spares_coll = db["spares_master"]

        # Check duplicate part no
        if spares_coll.find_one({"part_no": part_no}):
            return JsonResponse({"error": "Part number already exists"}, status=409)

        # Insert into DB
        spares_coll.insert_one({
            "part_no": part_no,
            "item_name": item_name,
            "no_of_bins": int(no_of_bins),
            "bin_nos": bin_nos,
            "rack_no": rack_no,
            "item_loc": item_loc,
            "created_by": user.get("username"),
            "created_at": datetime.now(ZoneInfo("Asia/Kolkata"))
        })

        return JsonResponse({"message": "Item added", "part_no": part_no}, status=201)

    except Exception as e:
        stack = traceback.format_exc()
        return JsonResponse({"error": str(e), "stack_trace": stack}, status=500)

@csrf_exempt
def spares_master_list(request):
    if request.method == "GET":
        try:
            part_no = request.GET.get("part_no")

            # If part_no is passed → return the single item's details
            if part_no:
                item = spares_master.find_one({"part_no": part_no}, {"_id": 0})
                if not item:
                    return JsonResponse({"error": "Item not found"}, status=404)
                return JsonResponse(item)

            # Otherwise → return full list
            items = list(spares_master.find({}, {"_id": 0}))
            return JsonResponse({"items": items})

        except Exception as e:
            print(traceback.format_exc())
            return JsonResponse({"error": str(e)}, status=500)

    return JsonResponse({"error": "Method Not Allowed"}, status=405)

@csrf_exempt
def spares_in(request):
    user = get_user_from_token(request)
    if request.method == "POST":
        try:
            data = json.loads(request.body.decode("utf-8"))
            part_no = data.get("part_no")
            recieved_from = data.get("recieved_from")
            qty_in = int(data.get("qty_in", 0))
            remarks = data.get("remarks")

            if not part_no or qty_in <= 0 or not recieved_from:
                return JsonResponse({"error": "Invalid data"}, status=400)

            # Find item in master list
            item = spares_master.find_one({"part_no": part_no})
            no_of_bins = spares_master.find_one({"part_no": part_no}).get("no_of_bins",0)
            bin_nos = spares_master.find_one({"part_no": part_no}).get("bin_nos",[])
            rack_no = spares_master.find_one({"part_no": part_no}).get("rack_no","")
            item_loc = spares_master.find_one({"part_no": part_no}).get("item_loc","")

            if not item:
                return JsonResponse({"error": "Item not found"}, status=404)

            # Current qty
            current_qty = int(item.get("qty", 0))

            # Date tracking
            entry_date = datetime.now(ZoneInfo("Asia/Kolkata"))
            date= entry_date

            # New quantity = old + incoming qty
            new_qty = current_qty + qty_in

            # Update master list
            spares_master.update_one(
                {"part_no": part_no},
                {
                    "$set": {"qty": new_qty},
                    "$push": {
                        "history": {
                            "type": "IN",
                            "qty": qty_in,
                            "recieved_from": recieved_from,
                            "date": entry_date,
                            "remarks": remarks,
                            "no_of_bins": no_of_bins,
                            "bin_nos": bin_nos,
                            "rack_no": rack_no,
                            "item_loc": item_loc
                        }
                    }
                }
            )

            # Log entry 
            spares_in_col.insert_one({
                "part_no": part_no,
                "qty_in": qty_in,
                "recieved_from": recieved_from,
                "previous_qty": current_qty,
                "new_qty": new_qty,
                "date": entry_date,
                "remarks": remarks,
                "no_of_bins": no_of_bins,
                "bin_nos": bin_nos,
                "rack_no": rack_no,
                "item_loc": item_loc
            })

            spares_audit.insert_one({
                "part_no": part_no,
                "date": datetime.now(tz=ZoneInfo("Asia/Kolkata")),
                "recieved_from": recieved_from,
                "in": qty_in,
                "out": 0,
                "qty_after": new_qty,
                "user": user,
                "remarks": remarks,
                "no_of_bins": no_of_bins,
                "bin_nos": bin_nos,
                "rack_no": rack_no,
                "item_loc": item_loc
            })

            return JsonResponse({"status": "success", "new_qty": new_qty})

        except Exception as e:
            print(traceback.format_exc())
            return JsonResponse({"error": str(e)}, status=500)

    return JsonResponse({"error": "Method Not Allowed"}, status=405)

@csrf_exempt
def spares_out(request):
    user = get_user_from_token(request)
    if request.method == "POST":
        try:
            data = json.loads(request.body.decode("utf-8"))
            part_no = data.get("part_no")
            qty_out = int(data.get("qty_out", 0))
            handing_over_to = data.get("handing_over_to")
            remarks= data.get("remarks")
            if not part_no or qty_out <= 0 or not handing_over_to:
                return JsonResponse({"error": "Invalid input"}, status=400)

            # Find item
            item = spares_master.find_one({"part_no": part_no})
            no_of_bins = spares_master.find_one({"part_no": part_no}).get("no_of_bins",0)
            bin_nos = spares_master.find_one({"part_no": part_no}).get("bin_nos",[])
            rack_no = spares_master.find_one({"part_no": part_no}).get("rack_no","")
            item_loc = spares_master.find_one({"part_no": part_no}).get("item_loc","")

            if not item:
                return JsonResponse({"error": "Item not found"}, status=404)

            current_qty = int(item.get("qty", 0))

            # Validate stock availability
            if qty_out > current_qty:
                return JsonResponse({"error": "Not enough stock"}, status=400)

            new_qty = current_qty - qty_out

            # Date
            from datetime import datetime
            from zoneinfo import ZoneInfo
            entry_date = datetime.now(ZoneInfo("Asia/Kolkata"))
            date= entry_date

            # Update master list
            spares_master.update_one(
                {"part_no": part_no},
                {
                    "$set": {"qty": new_qty},
                    "$push": {
                        "history": {
                            "type": "OUT",
                            "qty": qty_out,
                            "handed_to": handing_over_to,
                            "date": entry_date,
                            "remarks": remarks,
                            "no_of_bins": no_of_bins,
                            "bin_nos": bin_nos,
                            "rack_no": rack_no,
                            "item_loc": item_loc
                        }
                    }
                }
            )

            # Log outgoing
            spares_out_col.insert_one({
                "part_no": part_no,
                "qty_out": qty_out,
                "handing_over_to": handing_over_to,
                "previous_qty": current_qty,
                "new_qty": new_qty,
                "date": entry_date,
                "remarks": remarks,
                "no_of_bins": no_of_bins,
                "bin_nos": bin_nos,
                "rack_no": rack_no,
                "item_loc": item_loc
            })

            spares_audit.insert_one({
                "part_no": part_no,
                "date": datetime.now(tz=ZoneInfo("Asia/Kolkata")),
                "in": 0,
                "out": qty_out,
                "qty_after": new_qty,
                "user": user,
                "remarks": remarks,
                "handing_over_to": handing_over_to,
                "no_of_bins": no_of_bins,
                "bin_nos": bin_nos,
                "rack_no": rack_no,
                "item_loc": item_loc
            })

            return JsonResponse({"status": "success", "new_qty": new_qty})

        except Exception as e:
            print(traceback.format_exc())
            return JsonResponse({"error": str(e)}, status=500)

    return JsonResponse({"error": "Method Not Allowed"}, status=405)

@csrf_exempt
def spares_audit_view(request):
    if request.method == "GET":
        try:
            part_no = request.GET.get("part_no")

            audit_items = list(
                spares_audit.find({"part_no": part_no}, {"_id": 0}).sort("date", 1)
            )

            return JsonResponse({"audit": audit_items})

        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)

    return JsonResponse({"error": "Method Not Allowed"}, status=405)

@csrf_exempt
def spares_audit_filter(request):
    if request.method == "GET":
        try:
            part_no = request.GET.get("part_no")
            start_date = request.GET.get("start_date")
            end_date = request.GET.get("end_date")

            query = {"part_no": part_no}

            # If date filter applied
            if start_date and end_date:
                try:
                    start_dt = datetime.strptime(start_date, "%Y-%m-%d")

                    # end date should include entire day (23:59:59)
                    end_dt = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)

                    query["date"] = {
                        "$gte": start_dt,
                        "$lt": end_dt
                    }

                except Exception:
                    print("Date parsing error:", traceback.format_exc())
                    return JsonResponse({"error": "Invalid date format"}, status=400)

            audit_items = list(
                spares_audit.find(query, {"_id": 0}).sort("date", 1)
            )

            return JsonResponse({"audit": audit_items})

        except Exception as e:
            print("Error:", traceback.format_exc())
            return JsonResponse({"error": str(e)}, status=500)

    return JsonResponse({"error": "Method Not Allowed"}, status=405)

import re

def sort_key(part_no):
    part_no = str(part_no).strip()
    match = re.match(r'^(\d+)', part_no)

    if match:
        return (1, int(match.group(1)), part_no)
    else:
        return (2, float('inf'), part_no)

@csrf_exempt
def stock_check(request):
    if request.method == "GET":
        try:
            # Fetch stock same way as stock check
            items = list(spares_master.find({}, {"_id": 0}))
            items.sort(key=lambda x: sort_key(x.get("part_no", "")))

            # Build CSV
            import csv
            from io import StringIO

            output = StringIO()
            writer = csv.writer(output)

            # Header
            writer.writerow(["Sl No", "Part No","Item Name","Item Loc","Rack No","No of Bins","Bin No","Qty"])

            # Rows
            for idx, item in enumerate(items):
                writer.writerow([
                    idx + 1,
                    item.get("part_no", ""),
                    item.get("item_name", ""),
                    item.get("item_loc", ""),
                    item.get("rack_no", ""),
                    item.get("no_of_bins", 0),
                    item.get("bin_no", ""),
                    item.get("qty", 0)
                ])

            response = HttpResponse(output.getvalue(), content_type="text/csv")
            response["Content-Disposition"] = "attachment; filename=stock_report.csv"
            return response

        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)

    return JsonResponse({"error": "Method Not Allowed"}, status=405)
